from io import BytesIO, StringIO
import csv
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from contextlib import contextmanager
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@contextmanager
def get_db_connection(db_pool):
    """Get a database connection from the pool"""
    conn = None
    try:
        conn = db_pool.get_connection()
        yield conn
    except Exception as e:
        print(f"Database error: {str(e)}")
        raise
    finally:
        if conn:
            conn.close()


def generate_advanced_report(db_pool, start_date, end_date, report_id, format_type='pdf'):
    """Dispatcher for different report types"""
    try:
        if format_type == 'csv':
            return generate_csv_report(db_pool, start_date, end_date, report_id)
        elif format_type == 'excel':
            return generate_excel_report(db_pool, start_date, end_date, report_id)
        else:
            # PDF reports
            if report_id == 'sales_summary':
                return generate_sales_report(db_pool, start_date, end_date)
            elif report_id == 'revenue_analytics':
                return generate_revenue_report(db_pool, start_date, end_date)
            elif report_id == 'inventory_status':
                return generate_inventory_report(db_pool)
            elif report_id == 'customer_engagement':
                return generate_customer_report(db_pool, start_date, end_date)
            elif report_id == 'profit_loss':
                return generate_profit_loss_report(db_pool, start_date, end_date)
            elif report_id == 'balance_sheet':
                return generate_balance_sheet(db_pool, start_date, end_date)
            elif report_id == 'refund_return_summary':
                return generate_returns_report(db_pool, start_date, end_date)
            elif report_id == 'cancellation_report':
                return generate_cancellation_report(db_pool, start_date, end_date)
            elif report_id == 'tax_gst_compliance':
                return generate_gst_report(db_pool, start_date, end_date)
            else:
                return generate_sales_report(db_pool, start_date, end_date)
    except Exception as e:
        print(f"Error in advanced report generation: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, str(e)


# ======================== SALES SUMMARY ========================
def generate_sales_report(db_pool, start_date, end_date):
    """Comprehensive sales summary with daily breakdown, channels, and demographics"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            # 1. Gross vs Net Sales
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT o.id) as total_orders,
                    SUM(o.total_amount) as gross_sales,
                    SUM(CASE WHEN o.status = 'cancelled' THEN o.total_amount ELSE 0 END) as cancelled_sales,
                    SUM(CASE WHEN o.status = 'refunded' THEN o.return_amount ELSE 0 END) as refunded_sales,
                    SUM(CASE WHEN o.status NOT IN ('cancelled', 'refunded') THEN o.total_amount ELSE 0 END) as net_sales,
                    SUM(oi.quantity) as total_items
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
            """, (start_date, end_date))
            summary = cursor.fetchone()

            # 2. Sales Channel Breakdown
            cursor.execute("""
                SELECT 
                    COALESCE(sales_channel, 'Native') as channel,
                    COUNT(id) as order_count,
                    SUM(total_amount) as revenue
                FROM orders
                WHERE DATE(order_date) BETWEEN %s AND %s
                AND status != 'cancelled'
                GROUP BY channel
                ORDER BY revenue DESC
            """, (start_date, end_date))
            channels = cursor.fetchall()

            # 3. Category Performance
            cursor.execute("""
                SELECT 
                    c.name as category,
                    SUM(oi.quantity) as items_sold,
                    SUM(oi.price * oi.quantity) as revenue
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                LEFT JOIN categories c ON p.category = c.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
                GROUP BY p.category, c.name
                ORDER BY revenue DESC
            """, (start_date, end_date))
            categories = cursor.fetchall()

            # 4. Payment Method Breakdown
            cursor.execute("""
                SELECT 
                    payment_method,
                    COUNT(id) as order_count,
                    SUM(total_amount) as revenue
                FROM orders
                WHERE DATE(order_date) BETWEEN %s AND %s
                AND status != 'cancelled'
                GROUP BY payment_method
                ORDER BY revenue DESC
            """, (start_date, end_date))
            payments = cursor.fetchall()

            # 5. Geographic Demographics (City/State)
            cursor.execute("""
                SELECT 
                    JSON_UNQUOTE(JSON_EXTRACT(shipping_address, '$.state')) as state,
                    JSON_UNQUOTE(JSON_EXTRACT(shipping_address, '$.city')) as city,
                    COUNT(id) as order_count,
                    SUM(total_amount) as revenue
                FROM orders
                WHERE DATE(order_date) BETWEEN %s AND %s
                AND status != 'cancelled'
                AND shipping_address IS NOT NULL
                GROUP BY state, city
                ORDER BY revenue DESC
                LIMIT 10
            """, (start_date, end_date))
            demographics = cursor.fetchall()

            # Building PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=20)
            elements.append(Paragraph("ADVANCED SALES SUMMARY", title_style))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            if summary and summary['total_orders']:
                gross = summary['gross_sales'] or 0
                cancelled = summary['cancelled_sales'] or 0
                refunded = summary['refunded_sales'] or 0
                net = summary['net_sales'] or 0
                
                summary_data = [
                    ["Metric", "Amount (Rs.)"],
                    ["Total Orders (All statuses)", f"{summary['total_orders']}"],
                    ["Gross Sales", f"{gross:,.2f}"],
                    ["Less: Cancellations", f"({cancelled:,.2f})"],
                    ["Less: Refunds", f"({refunded:,.2f})"],
                    ["NET SALES", f"{net:,.2f}"]
                ]
                summary_table = Table(summary_data, colWidths=[250, 150])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
                    ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold')
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))

            # Helper function for generic tables
            def add_table(title, headers, data_rows, col_widths, align_right_cols):
                if not data_rows: return
                elements.append(Paragraph(title, styles['Heading3']))
                table_data = [headers]
                for row in data_rows:
                    table_data.append(row)
                t = Table(table_data, colWidths=col_widths)
                style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ]
                for col in align_right_cols:
                    style.append(('ALIGN', (col, 0), (col, -1), 'RIGHT'))
                t.setStyle(TableStyle(style))
                elements.append(t)
                elements.append(Spacer(1, 15))

            # Channels
            channel_data = [[i+1, c['channel'], c['order_count'], f"{c['revenue'] or 0:,.2f}"] for i, c in enumerate(channels)]
            add_table("Sales by Channel", ["Sr.", "Channel", "Orders", "Revenue (Rs.)"], channel_data, [30, 150, 100, 120], [2, 3])

            # Categories
            cat_data = [[i+1, c['category'] or 'Uncategorized', c['items_sold'], f"{c['revenue'] or 0:,.2f}"] for i, c in enumerate(categories)]
            add_table("Category Performance", ["Sr.", "Category", "Items Sold", "Revenue (Rs.)"], cat_data, [30, 150, 100, 120], [2, 3])

            # Payments
            pay_data = [[i+1, p['payment_method'] or 'Unknown', p['order_count'], f"{p['revenue'] or 0:,.2f}"] for i, p in enumerate(payments)]
            add_table("Payment Methods", ["Sr.", "Payment Gateway", "Orders", "Revenue (Rs.)"], pay_data, [30, 150, 100, 120], [2, 3])

            # Demographics
            demo_data = [[i+1, f"{d['state']}", d['order_count'], f"{d['revenue'] or 0:,.2f}"] for i, d in enumerate(demographics)]
            add_table("Top Regional Markets", ["Sr.", "Location", "Orders", "Revenue (Rs.)"], demo_data, [30, 250, 80, 120], [2, 3])

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)





# ======================== REVENUE ANALYTICS ========================
def generate_revenue_report(db_pool, start_date, end_date):
    """Revenue analytics: gross revenue, taxes, shipping collections"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            cursor.execute("""
                SELECT COALESCE(SUM(o.total_amount), 0) as gross_revenue
                FROM orders o
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
            """, (start_date, end_date))
            gross = cursor.fetchone()['gross_revenue']

            cursor.execute("""
                SELECT COALESCE(SUM((oi.price * oi.quantity) * (p.gst_rate / 100)), 0) as total_tax
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
            """, (start_date, end_date))
            tax = cursor.fetchone()['total_tax']

            try:
                cursor.execute("""
                    SELECT COALESCE(SUM(o.shipping_charge), 0) as shipping
                    FROM orders o
                    WHERE DATE(o.order_date) BETWEEN %s AND %s
                    AND o.status != 'cancelled'
                """, (start_date, end_date))
                shipping = cursor.fetchone()['shipping']
            except:
                shipping = 0

            net_revenue = gross - tax

            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.order_date, '%Y-%m') as month,
                    COUNT(DISTINCT o.id) as order_count,
                    COALESCE(SUM(o.total_amount), 0) as monthly_revenue,
                    COALESCE(SUM((oi.price * oi.quantity) * (p.gst_rate / 100)), 0) as monthly_tax
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
                GROUP BY DATE_FORMAT(o.order_date, '%Y-%m')
                ORDER BY month
            """, (start_date, end_date))
            monthly = cursor.fetchall()

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("REVENUE ANALYTICS", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            rev_data = [
                ["Metric", "Amount (Rs.)"],
                ["Gross Revenue", f"{gross:,.2f}"],
                ["Total Tax (GST)", f"{tax:,.2f}"],
                ["Shipping Collections", f"{shipping:,.2f}"],
                ["Net Revenue (after tax)", f"{net_revenue:,.2f}"]
            ]
            rev_table = Table(rev_data, colWidths=[200, 200])
            rev_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(rev_table)
            elements.append(Spacer(1, 20))

            if monthly:
                elements.append(Paragraph("Monthly Breakdown", styles['Heading3']))
                monthly_data = [["Sr.", "Month", "Orders", "Revenue (Rs.)", "Tax (Rs.)"]]
                for i, m in enumerate(monthly, 1):
                    monthly_data.append([
                        i,
                        m['month'],
                        m['order_count'],
                        f"{m['monthly_revenue']:,.2f}",
                        f"{m['monthly_tax']:,.2f}"
                    ])
                monthly_table = Table(monthly_data, colWidths=[30, 90, 80, 130, 130])
                monthly_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ]))
                elements.append(monthly_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== INVENTORY STATUS ========================
def generate_inventory_report(db_pool):
    """Current stock levels, valuation, reorder alerts"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT 
                    p.sku,
                    p.name,
                    c.name as category,
                    p.stock_quantity,
                    p.reorder_level,
                    p.price as unit_price,
                    (p.stock_quantity * p.price) as stock_value,
                    CASE 
                        WHEN p.stock_quantity <= 0 THEN 'Out of Stock'
                        WHEN p.stock_quantity <= p.reorder_level THEN 'Low Stock'
                        ELSE 'In Stock'
                    END as stock_status
                FROM products p 
                LEFT JOIN categories c ON p.category = c.id
                ORDER BY p.stock_quantity ASC
            """)
            items = cursor.fetchall()

            total_value = sum(float(item['stock_value'] or 0) for item in items)
            low_stock_items = [i for i in items if i['stock_status'] == 'Low Stock']
            out_of_stock = [i for i in items if i['stock_status'] == 'Out of Stock']

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("INVENTORY STATUS", styles['Title']))
            elements.append(Spacer(1, 10))

            summary_data = [
                ["Total SKUs", len(items)],
                ["Total Inventory Value (Rs.)", f"{total_value:,.2f}"],
                ["Low Stock Items", len(low_stock_items)],
                ["Out of Stock Items", len(out_of_stock)]
            ]
            sum_table = Table(summary_data, colWidths=[180, 180])
            sum_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(sum_table)
            elements.append(Spacer(1, 20))

            if items:
                inv_data = [["Sr.", "SKU", "Product", "Category", "Stock", "Reorder", "Value (Rs.)", "Status"]]
                for i, item in enumerate(items, 1):
                    inv_data.append([
                        i,
                        item['sku'],
                        Paragraph(item['name'], styles['Normal']),
                        Paragraph(item['category'] or 'Uncategorized', styles['Normal']),
                        item['stock_quantity'],
                        item['reorder_level'],
                        f"{item['stock_value']:,.2f}",
                        item['stock_status']
                    ])
                inv_table = Table(inv_data, colWidths=[30, 80, 170, 90, 60, 60, 90, 90])
                inv_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (2, 0), (3, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (4, 0), (6, -1), 'RIGHT'),
                ]))
                elements.append(inv_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== CUSTOMER ENGAGEMENT ========================
def generate_customer_report(db_pool, start_date, end_date):
    """New signups, active users, CLV metrics"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            cursor.execute("""
                SELECT COUNT(*) as new_customers
                FROM users
                WHERE DATE(created_at) BETWEEN %s AND %s
            """, (start_date, end_date))
            new_count = cursor.fetchone()['new_customers']

            cursor.execute("""
                SELECT COUNT(DISTINCT user_id) as active_users
                FROM orders
                WHERE DATE(order_date) BETWEEN %s AND %s
                AND status != 'cancelled'
            """, (start_date, end_date))
            active = cursor.fetchone()['active_users'] or 0

            cursor.execute("""
                SELECT 
                    u.id,
                    u.username,
                    u.email,
                    COUNT(o.id) as order_count,
                    COALESCE(SUM(o.total_amount), 0) as lifetime_value
                FROM users u
                LEFT JOIN orders o ON u.id = o.user_id AND o.status != 'cancelled'
                GROUP BY u.id
                ORDER BY lifetime_value DESC
                LIMIT 20
            """)
            customers = cursor.fetchall()
            avg_clv = sum(c['lifetime_value'] for c in customers) / len(customers) if customers else 0

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("CUSTOMER ENGAGEMENT", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            kpi_data = [
                ["Metric", "Value"],
                ["New Signups (this period)", new_count],
                ["Active Users (placed orders)", active],
                ["Average Customer Lifetime Value (CLV) (Rs.)", f"{avg_clv:,.2f}"]
            ]
            kpi_table = Table(kpi_data, colWidths=[280, 150])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(kpi_table)
            elements.append(Spacer(1, 20))

            if customers:
                elements.append(Paragraph("Top Customers by Lifetime Value", styles['Heading3']))
                cust_data = [["Sr.", "Username", "Email", "Orders", "Total Spent (Rs.)"]]
                for i, c in enumerate(customers[:15], 1):
                    cust_data.append([
                        i,
                        Paragraph(c['username'] or 'Unknown', styles['Normal']),
                        Paragraph(c['email'] or 'Unknown', styles['Normal']),
                        c['order_count'],
                        f"{c['lifetime_value']:,.2f}"
                    ])
                cust_table = Table(cust_data, colWidths=[30, 110, 180, 70, 110])
                cust_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (2, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                ]))
                elements.append(cust_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== PROFIT & LOSS ========================
def generate_profit_loss_report(db_pool, start_date, end_date):
    """Revenue vs material costs and operational overheads"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            cursor.execute("""
                SELECT SUM(total_amount) as revenue
                FROM orders
                WHERE DATE(order_date) BETWEEN %s AND %s
                AND status != 'cancelled'
            """, (start_date, end_date))
            revenue = float(cursor.fetchone()['revenue'] or 0)

            cursor.execute("""
                SELECT SUM(oi.quantity * COALESCE(p.material_cost, 0)) as material_cost
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
            """, (start_date, end_date))
            material_cost = float(cursor.fetchone()['material_cost'] or 0)

            gross_profit = revenue - material_cost
            gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0

            overhead_percent = 8
            overhead = revenue * (overhead_percent / 100)
            net_profit = gross_profit - overhead
            net_margin = (net_profit / revenue * 100) if revenue > 0 else 0

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("PROFIT & LOSS STATEMENT", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            def format_currency(val):
                return f"({abs(val):,.2f})" if val < 0 else f"{val:,.2f}"

            gross_label = "Gross Profit" if gross_profit >= 0 else "Gross Loss"
            net_label = "Net Profit" if net_profit >= 0 else "Net Loss"

            pl_data = [
                ["Particulars", "Amount (Rs.)"],
                ["Total Revenue", format_currency(revenue)],
                ["Less: Material Cost", f"({material_cost:,.2f})"],
                [gross_label, format_currency(gross_profit)],
                [f"Less: Operational Overhead ({overhead_percent}%)", f"({overhead:,.2f})"],
                [net_label, format_currency(net_profit)],
                ["Gross Margin %", f"{gross_margin:.1f}%"],
                ["Net Margin %", f"{net_margin:.1f}%"]
            ]
            pl_table = Table(pl_data, colWidths=[250, 150])
            pl_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 6), (-1, 7), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ]))
            elements.append(pl_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== BALANCE SHEET ========================
def generate_balance_sheet(db_pool, start_date, end_date):
    """Snapshot of assets (inventory, receivables) and liabilities"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            cursor.execute("""
                SELECT COALESCE(SUM(stock_quantity * price), 0) as inventory_value
                FROM products
            """)
            inventory_val = cursor.fetchone()['inventory_value']

            cursor.execute("""
                SELECT COALESCE(SUM(total_amount), 0) as receivables
                FROM orders
                WHERE status NOT IN ('cancelled', 'delivered', 'refunded')
            """)
            receivables = cursor.fetchone()['receivables']

            cursor.execute("""
                SELECT COALESCE(SUM(total_amount), 0) as cash_received
                FROM orders
                WHERE status IN ('delivered', 'completed')
            """)
            cash = cursor.fetchone()['cash_received']

            total_assets = inventory_val + receivables + cash

            cursor.execute("""
                SELECT COALESCE(SUM((oi.price * oi.quantity) * (p.gst_rate / 100)), 0) as gst_payable
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
            """, (start_date, end_date))
            gst_payable = cursor.fetchone()['gst_payable']

            cursor.execute("""
                SELECT COALESCE(SUM(return_amount), 0) as refunds_pending
                FROM orders
                WHERE status = 'refunded' AND returned_at IS NULL
            """)
            refunds_pending = cursor.fetchone()['refunds_pending']

            total_liabilities = gst_payable + refunds_pending
            equity = total_assets - total_liabilities

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("BALANCE SHEET", styles['Title']))
            elements.append(Paragraph(f"As on {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            assets_data = [
                ["ASSETS", "Amount (Rs.)"],
                ["Inventory Value", f"{inventory_val:,.2f}"],
                ["Accounts Receivable", f"{receivables:,.2f}"],
                ["Cash & Bank", f"{cash:,.2f}"],
                ["Total Assets", f"{total_assets:,.2f}"]
            ]
            assets_table = Table(assets_data, colWidths=[250, 150])
            assets_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 4), (-1, 4), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ]))
            elements.append(assets_table)
            elements.append(Spacer(1, 20))

            liab_data = [
                ["LIABILITIES & EQUITY", "Amount (Rs.)"],
                ["GST Payable", f"{gst_payable:,.2f}"],
                ["Refunds Obligation", f"{refunds_pending:,.2f}"],
                ["Total Liabilities", f"{total_liabilities:,.2f}"],
                ["Owner's Equity", f"{equity:,.2f}"],
                ["Total Liabilities & Equity", f"{total_assets:,.2f}"]
            ]
            liab_table = Table(liab_data, colWidths=[250, 150])
            liab_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BACKGROUND', (0, 3), (-1, 3), colors.lightgrey),
                ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ]))
            elements.append(liab_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== REFUNDS & RETURNS ========================
def generate_returns_report(db_pool, start_date, end_date):
    """Detailed logs of return requests and approved refunds"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT 
                    o.id as order_id,
                    r.requested_date as return_date,
                    r.status,
                    o.return_amount as refund_amount,
                    r.reason,
                    r.remarks
                FROM order_returns r
                JOIN orders o ON r.order_id = o.id
                WHERE DATE(r.requested_date) BETWEEN %s AND %s
                ORDER BY r.requested_date DESC
            """, (start_date, end_date))
            returns = cursor.fetchall()

            total_refunded = sum(float(r['refund_amount'] or 0) for r in returns if r['status'] == 'completed')
            pending_returns = len([r for r in returns if r['status'] in ('requested', 'approved', 'processing')])

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("REFUNDS & RETURNS SUMMARY", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            summary_data = [
                ["Total Refunded Amount (Rs.)", f"{total_refunded:,.2f}"],
                ["Pending Returns", pending_returns],
                ["Total Return Requests", len(returns)]
            ]
            sum_table = Table([list(summary_data[0]), list(summary_data[1]), list(summary_data[2])], colWidths=[200, 150])
            sum_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(sum_table)
            elements.append(Spacer(1, 20))

            if returns:
                return_data = [["Sr.", "Order ID", "Return Date", "Status", "Refund Amount (Rs.)", "Reason"]]
                for i, r in enumerate(returns, 1):
                    return_data.append([
                        i,
                        f"#{r['order_id']}",
                        r['return_date'].strftime('%d-%b-%Y'),
                        r['status'].capitalize(),
                        f"{(r['refund_amount'] or 0):,.2f}",
                        (r['reason'] or '')[:50]
                    ])
                ret_table = Table(return_data, colWidths=[30, 70, 90, 80, 110, 180])
                ret_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                ]))
                elements.append(ret_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== CANCELLATION REPORT ========================
def generate_cancellation_report(db_pool, start_date, end_date):
    """Order cancellations and lost revenue opportunities"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT 
                    id as order_id,
                    total_amount,
                    cancelled_at,
                    cancellation_reason
                FROM orders 
                WHERE status = 'cancelled' 
                AND DATE(cancelled_at) BETWEEN %s AND %s
                ORDER BY cancelled_at DESC
            """, (start_date, end_date))
            cancellations = cursor.fetchall()

            total_lost = sum(float(c['total_amount'] or 0) for c in cancellations)

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("ORDER CANCELLATION REPORT", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            summary_data = [
                ["Total Cancelled Orders", len(cancellations)],
                ["Total Lost Revenue (Rs.)", f"{total_lost:,.2f}"]
            ]
            sum_table = Table(summary_data, colWidths=[200, 200])
            sum_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(sum_table)
            elements.append(Spacer(1, 20))

            if cancellations:
                cancel_data = [["Sr.", "Order ID", "Amount (Rs.)", "Cancelled At", "Reason"]]
                for i, c in enumerate(cancellations, 1):
                    cancel_data.append([
                        i,
                        f"#{c['order_id']}",
                        f"{c['total_amount']:,.2f}",
                        c['cancelled_at'].strftime('%d-%b-%Y %H:%M') if c['cancelled_at'] else 'N/A',
                        (c['cancellation_reason'] or 'Not specified')[:60]
                    ])
                cancel_table = Table(cancel_data, colWidths=[30, 80, 100, 140, 220])
                cancel_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ]))
                elements.append(cancel_table)

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)


# ======================== GST COMPLIANCE ========================
def generate_gst_report(db_pool, start_date, end_date):
    """HSN-wise tax breakdown for GSTR filing"""
    try:
        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT 
                    p.hsn_code,
                    p.gst_rate,
                    SUM(oi.quantity) - COALESCE(SUM(ri.quantity), 0) as quantity,
                    SUM(oi.taxable_value) - COALESCE(SUM(ri.refund_taxable_value), 0) as taxable_value,
                    (SUM(oi.cgst_amount) + SUM(oi.sgst_amount) + SUM(oi.igst_amount)) - COALESCE(SUM(ri.refund_cgst) + SUM(ri.refund_sgst) + SUM(ri.refund_igst), 0) as total_gst
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                LEFT JOIN (
                    SELECT ri.product_id, orr.order_id, 
                           SUM(ri.quantity) as quantity,
                           SUM(ri.refund_taxable_value) as refund_taxable_value,
                           SUM(ri.refund_cgst) as refund_cgst,
                           SUM(ri.refund_sgst) as refund_sgst,
                           SUM(ri.refund_igst) as refund_igst
                    FROM return_items ri
                    JOIN order_returns orr ON ri.return_id = orr.id
                    WHERE orr.status = 'completed'
                    GROUP BY ri.product_id, orr.order_id
                ) ri ON ri.product_id = oi.product_id AND ri.order_id = oi.order_id
                WHERE o.status != 'cancelled' 
                AND DATE(o.order_date) BETWEEN %s AND %s
                GROUP BY p.hsn_code, p.gst_rate
                ORDER BY p.hsn_code
            """, (start_date, end_date))
            rows = cursor.fetchall()

            total_taxable = 0
            total_gst = 0

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("GST COMPLIANCE REPORT (HSN-WISE)", styles['Title']))
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Heading2']))
            elements.append(Spacer(1, 20))

            if rows:
                gst_data = [["Sr.", "HSN Code", "GST Rate", "Qty", "Taxable (Rs.)", "CGST (Rs.)", "SGST (Rs.)", "Total GST (Rs.)"]]
                for i, r in enumerate(rows, 1):
                    taxable = float(r['taxable_value'] or 0)
                    gst = float(r['total_gst'] or 0)
                    total_taxable += taxable
                    total_gst += gst
                    gst_data.append([
                        i,
                        r['hsn_code'] or 'N/A',
                        f"{r['gst_rate']}%",
                        r['quantity'],
                        f"{taxable:,.2f}",
                        f"{gst/2:,.2f}",
                        f"{gst/2:,.2f}",
                        f"{gst:,.2f}"
                    ])

                gst_table = Table(gst_data, colWidths=[30, 90, 70, 60, 110, 90, 90, 100])
                gst_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                ]))
                elements.append(gst_table)
                elements.append(Spacer(1, 20))

                elements.append(Paragraph(f"Total Taxable Value: Rs.{total_taxable:,.2f}", styles['Normal']))
                elements.append(Paragraph(f"Total GST Collected: Rs.{total_gst:,.2f}", styles['Normal']))
                elements.append(Paragraph(f"CGST: Rs.{total_gst/2:,.2f} | SGST: Rs.{total_gst/2:,.2f}", styles['Normal']))
            else:
                elements.append(Paragraph("No taxable transactions found in the selected period.", styles['Normal']))

            doc.build(elements)
            buffer.seek(0)
            return buffer, None

    except Exception as e:
        return None, str(e)

# ======================== Excel EXPORTS ========================
def generate_excel_report(db_pool, start_date, end_date, report_id):
    """Generate Excel (.xlsx) report for any report type"""
    wb = Workbook()
    ws = wb.active
    ws.title = report_id.replace('_', ' ').title()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    with get_db_connection(db_pool) as conn:
        cursor = conn.cursor(dictionary=True)

        if report_id == 'sales_summary':
            # Overall summary
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT o.id) as total_orders,
                    SUM(o.total_amount) as total_sales,
                    SUM(oi.quantity) as total_items,
                    AVG(o.total_amount) as avg_order_value
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
            """, (start_date, end_date))
            summary = cursor.fetchone()

            # Daily breakdown
            cursor.execute("""
                SELECT 
                    DATE(o.order_date) as order_date,
                    COUNT(DISTINCT o.id) as order_count,
                    SUM(o.total_amount) as daily_sales,
                    SUM(oi.quantity) as items_sold
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                WHERE DATE(o.order_date) BETWEEN %s AND %s
                AND o.status != 'cancelled'
                GROUP BY DATE(o.order_date)
                ORDER BY DATE(o.order_date)
            """, (start_date, end_date))
            daily = cursor.fetchall()

            # Top products
            cursor.execute("""
                SELECT p.name, SUM(oi.quantity) as qty, SUM(oi.price * oi.quantity) as revenue
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s AND o.status != 'cancelled'
                GROUP BY p.id ORDER BY revenue DESC LIMIT 10
            """, (start_date, end_date))
            top = cursor.fetchall()

            # Write summary
            # (Titles removed as requested)
            ws.append(["Metric", "Value"])
            ws.append(["Total Orders", summary['total_orders']])
            ws.append(["Total Sales (Rs.)", f"{summary['total_sales']:,.2f}"])
            ws.append(["Total Items Sold", summary['total_items'] or 0])
            ws.append(["Average Order Value (Rs.)", f"{summary['avg_order_value']:,.2f}"])
            ws.append([])
            ws.append(["Sr.", "Date", "Orders", "Revenue (Rs.)", "Items Sold"])
            for i, row in enumerate(daily, 1):
                ws.append([i, row['order_date'].strftime('%Y-%m-%d'), row['order_count'], f"{row['daily_sales']:,.2f}", row['items_sold'] or 0])
            ws.append([])
            ws.append(["Sr.", "Product Name", "Quantity Sold", "Revenue (Rs.)"])
            for i, row in enumerate(top, 1):
                ws.append([i, row['name'], row['qty'], f"{row['revenue']:,.2f}"])

        elif report_id == 'inventory_status':
            cursor.execute("""
                SELECT p.sku, p.name, c.name as category, p.stock_quantity, p.reorder_level, p.price,
                       (p.stock_quantity * p.price) as stock_value,
                       CASE WHEN p.stock_quantity <= 0 THEN 'Out of Stock'
                            WHEN p.stock_quantity <= p.reorder_level THEN 'Low Stock'
                            ELSE 'In Stock' END as status
                FROM products p
                LEFT JOIN categories c ON p.category = c.id
                ORDER BY p.stock_quantity ASC
            """)
            items = cursor.fetchall()
            # ws.append(["INVENTORY STATUS"])
            # ws.append([])
            ws.append(["Sr.", "SKU", "Product", "Category", "Stock", "Reorder", "Unit Price (Rs.)", "Stock Value (Rs.)", "Status"])
            for i, row in enumerate(items, 1):
                ws.append([i, row['sku'], row['name'], row['category'], row['stock_quantity'],
                           row['reorder_level'], f"{row['price']:.2f}", f"{row['stock_value']:.2f}", row['status']])

        elif report_id == 'customer_engagement':
            cursor.execute("""
                SELECT u.username, u.email, u.created_at,
                       COUNT(o.id) as orders, COALESCE(SUM(o.total_amount),0) as spent
                FROM users u
                LEFT JOIN orders o ON u.id = o.user_id AND o.status != 'cancelled'
                GROUP BY u.id ORDER BY spent DESC
            """)
            customers = cursor.fetchall()
            # ws.append(["CUSTOMER ENGAGEMENT"])
            # ws.append([f"Period: {start_date} to {end_date}"])
            # ws.append([])
            ws.append(["Sr.", "Username", "Email", "Joined", "Orders", "Total Spent (Rs.)"])
            for i, row in enumerate(customers, 1):
                ws.append([i, row['username'], row['email'], row['created_at'].strftime('%Y-%m-%d'), row['orders'], f"{row['spent']:.2f}"])

        elif report_id == 'cancellation_report':
            cursor.execute("""
                SELECT id, total_amount, cancelled_at, cancellation_reason
                FROM orders WHERE status = 'cancelled' AND DATE(cancelled_at) BETWEEN %s AND %s
                ORDER BY cancelled_at DESC
            """, (start_date, end_date))
            rows = cursor.fetchall()
            # ws.append(["CANCELLATION REPORT"])
            # ws.append([f"Period: {start_date} to {end_date}"])
            # ws.append([])
            ws.append(["Sr.", "Order ID", "Amount (Rs.)", "Cancelled At", "Reason"])
            for i, row in enumerate(rows, 1):
                ws.append([i, row['id'], f"{row['total_amount']:.2f}", row['cancelled_at'].strftime('%Y-%m-%d %H:%M'), row['cancellation_reason']])

        elif report_id == 'refund_return_summary':
            cursor.execute("""
                SELECT o.id, r.requested_date, r.status, o.return_amount, r.reason
                FROM order_returns r JOIN orders o ON r.order_id = o.id
                WHERE DATE(r.requested_date) BETWEEN %s AND %s
                ORDER BY r.requested_date DESC
            """, (start_date, end_date))
            rows = cursor.fetchall()
            # ws.append(["REFUNDS & RETURNS"])
            # ws.append([f"Period: {start_date} to {end_date}"])
            # ws.append([])
            ws.append(["Sr.", "Order ID", "Return Date", "Status", "Refund Amount (Rs.)", "Reason"])
            for i, row in enumerate(rows, 1):
                ws.append([i, row['id'], row['requested_date'].strftime('%Y-%m-%d'), row['status'].capitalize(),
                           f"{row['return_amount'] or 0:.2f}", row['reason'][:100]])

        elif report_id == 'tax_gst_compliance':
            cursor.execute("""
                SELECT 
                    p.hsn_code, p.gst_rate,
                    SUM(oi.quantity) - COALESCE(SUM(ri.quantity), 0) as qty,
                    SUM(oi.taxable_value) - COALESCE(SUM(ri.refund_taxable_value), 0) as taxable,
                    (SUM(oi.cgst_amount) + SUM(oi.sgst_amount) + SUM(oi.igst_amount)) - COALESCE(SUM(ri.refund_cgst) + SUM(ri.refund_sgst) + SUM(ri.refund_igst), 0) as gst
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                LEFT JOIN (
                    SELECT ri.product_id, orr.order_id, 
                           SUM(ri.quantity) as quantity, SUM(ri.refund_taxable_value) as refund_taxable_value,
                           SUM(ri.refund_cgst) as refund_cgst, SUM(ri.refund_sgst) as refund_sgst, SUM(ri.refund_igst) as refund_igst
                    FROM return_items ri JOIN order_returns orr ON ri.return_id = orr.id
                    WHERE orr.status = 'completed' GROUP BY ri.product_id, orr.order_id
                ) ri ON ri.product_id = oi.product_id AND ri.order_id = oi.order_id
                WHERE o.status != 'cancelled' AND DATE(o.order_date) BETWEEN %s AND %s
                GROUP BY p.hsn_code, p.gst_rate
            """, (start_date, end_date))
            rows = cursor.fetchall()
            # ws.append(["GST COMPLIANCE (HSN-WISE)"])
            # ws.append([f"Period: {start_date} to {end_date}"])
            # ws.append([])
            ws.append(["Sr.", "HSN Code", "GST Rate", "Quantity", "Taxable Value (Rs.)", "GST Amount (Rs.)"])
            for row in rows:
                ws.append([row['hsn_code'], f"{row['gst_rate']}%", row['qty'], f"{row['taxable']:.2f}", f"{row['gst']:.2f}"])

        elif report_id == 'profit_loss':
            # Simplified P&L
            cursor.execute("SELECT SUM(total_amount) as revenue FROM orders WHERE DATE(order_date) BETWEEN %s AND %s AND status != 'cancelled'", (start_date, end_date))
            revenue = float(cursor.fetchone()['revenue'] or 0)
            cursor.execute("""
                SELECT SUM(oi.quantity * COALESCE(p.material_cost,0)) as material_cost
                FROM order_items oi JOIN products p ON oi.product_id = p.id
                JOIN orders o ON oi.order_id = o.id
                WHERE DATE(o.order_date) BETWEEN %s AND %s AND o.status != 'cancelled'
            """, (start_date, end_date))
            material = float(cursor.fetchone()['material_cost'] or 0)
            gross = revenue - material
            overhead = revenue * 0.08
            net = gross - overhead
            # ws.append(["PROFIT & LOSS STATEMENT"])
            # ws.append([f"Period: {start_date} to {end_date}"])
            # ws.append([])
            ws.append(["Particulars", "Amount (Rs.)"])
            ws.append(["Total Revenue", f"{revenue:,.2f}"])
            ws.append(["Less: Material Cost", f"({material:,.2f})"])
            ws.append(["Gross Profit", f"{gross:,.2f}"])
            ws.append(["Less: Operational Overhead (8%)", f"({overhead:,.2f})"])
            ws.append(["Net Profit", f"{net:,.2f}"])

        elif report_id == 'balance_sheet':
            # Simplified Balance Sheet
            cursor.execute("SELECT COALESCE(SUM(stock_quantity * price),0) as inventory FROM products")
            inv = cursor.fetchone()['inventory']
            cursor.execute("SELECT COALESCE(SUM(total_amount),0) as receivables FROM orders WHERE status NOT IN ('cancelled','delivered','refunded')")
            rec = cursor.fetchone()['receivables']
            cursor.execute("SELECT COALESCE(SUM(total_amount),0) as cash FROM orders WHERE status IN ('delivered','completed')")
            cash = cursor.fetchone()['cash']
            total_assets = inv + rec + cash
            # ws.append(["BALANCE SHEET"])
            # ws.append([f"As on {end_date}"])
            # ws.append([])
            ws.append(["ASSETS", "Amount (Rs.)"])
            ws.append(["Inventory Value", f"{inv:,.2f}"])
            ws.append(["Accounts Receivable", f"{rec:,.2f}"])
            ws.append(["Cash & Bank", f"{cash:,.2f}"])
            ws.append(["Total Assets", f"{total_assets:,.2f}"])
            ws.append([])
            ws.append(["LIABILITIES & EQUITY", "Amount (Rs.)"])
            ws.append(["GST Payable (estimated)", "0.00"])  # Can compute if needed
            ws.append(["Owner's Equity", f"{total_assets:,.2f}"])
            ws.append(["Total Liabilities & Equity", f"{total_assets:,.2f}"])

        else:
            # fallback: sales summary
            cursor.execute("SELECT id, total_amount, order_date FROM orders WHERE DATE(order_date) BETWEEN %s AND %s", (start_date, end_date))
            rows = cursor.fetchall()
            # ws.append(["Orders"])
            ws.append(["Order ID", "Date", "Amount (Rs.)"])
            for row in rows:
                ws.append([row['id'], row['order_date'].strftime('%Y-%m-%d'), f"{row['total_amount']:.2f}"])

    # Apply header styling
    for row in range(1, ws.max_row+1):
        first_cell = ws.cell(row=row, column=1).value
        # Logic to identify header rows: Title rows (row 1 or empty row before) or rows containing "Sr."/ "Metric" / etc.
        is_header = False
        if row == 1:
            is_header = True
        elif first_cell:
            header_keywords = ["Sr.", "Metric", "Particulars", "ASSETS", "LIABILITIES", "Order ID", "Username", "SKU", "HSN Code"]
            if any(key in str(first_cell) for key in header_keywords):
                is_header = True
        
        if is_header:
            for cell in ws[row]:
                if isinstance(cell, openpyxl.cell.cell.Cell):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

    # Auto-size columns (simple)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None

# ======================== CSV EXPORTS ========================
def generate_csv_report(db_pool, start_date, end_date, report_id):
    """CSV export for various report types"""
    try:
        output = StringIO()
        writer = csv.writer(output)

        with get_db_connection(db_pool) as conn:
            cursor = conn.cursor(dictionary=True)

            if report_id == 'sales_summary':
                writer.writerow(["Sr.", "Order ID", "Date", "Amount (Rs.)", "Status"])
                cursor.execute("""
                    SELECT id, order_date, total_amount, status
                    FROM orders
                    WHERE DATE(order_date) BETWEEN %s AND %s
                    ORDER BY order_date DESC
                """, (start_date, end_date))
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['id'], row['order_date'], f"{row['total_amount']:.2f}", row['status']])

            elif report_id == 'inventory_status':
                writer.writerow(["Sr.", "SKU", "Product", "Category", "Stock", "Reorder Level", "Unit Price (Rs.)", "Stock Value (Rs.)"])
                cursor.execute("""
                    SELECT p.sku, p.name, c.name as category, p.stock_quantity, p.reorder_level, p.price,
                           (p.stock_quantity * p.price) as stock_value
                    FROM products p
                    LEFT JOIN categories c ON p.category = c.id
                    ORDER BY p.stock_quantity ASC
                """)
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['sku'], row['name'], row['category'], row['stock_quantity'],
                                     row['reorder_level'], f"{row['price']:.2f}", f"{row['stock_value']:.2f}"])

            elif report_id == 'customer_engagement':
                writer.writerow(["Sr.", "User ID", "Username", "Email", "Joined", "Total Orders", "Total Spent (Rs.)"])
                cursor.execute("""
                    SELECT u.id, u.username, u.email, u.created_at,
                           COUNT(o.id) as orders, COALESCE(SUM(o.total_amount),0) as spent
                    FROM users u
                    LEFT JOIN orders o ON u.id = o.user_id AND o.status != 'cancelled'
                    GROUP BY u.id
                    ORDER BY spent DESC
                """)
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['id'], row['username'], row['email'], row['created_at'],
                                     row['orders'], f"{row['spent']:.2f}"])

            elif report_id == 'cancellation_report':
                writer.writerow(["Sr.", "Order ID", "Amount (Rs.)", "Cancelled At", "Reason"])
                cursor.execute("""
                    SELECT id, total_amount, cancelled_at, cancellation_reason
                    FROM orders
                    WHERE status = 'cancelled' AND DATE(cancelled_at) BETWEEN %s AND %s
                """, (start_date, end_date))
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['id'], f"{row['total_amount']:.2f}", row['cancelled_at'], row['cancellation_reason']])

            elif report_id == 'refund_return_summary':
                writer.writerow(["Sr.", "Order ID", "Return Date", "Status", "Refund Amount (Rs.)", "Reason"])
                cursor.execute("""
                    SELECT o.id, r.requested_date, r.status, o.return_amount, r.reason
                    FROM order_returns r
                    JOIN orders o ON r.order_id = o.id
                    WHERE DATE(r.requested_date) BETWEEN %s AND %s
                """, (start_date, end_date))
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['id'], row['requested_date'], row['status'],
                                     f"{row['return_amount'] or 0:.2f}", row['reason']])

            elif report_id == 'tax_gst_compliance':
                writer.writerow(["Sr.", "HSN Code", "GST Rate", "Taxable Value (Rs.)", "GST Amount (Rs.)"])
                cursor.execute("""
                    SELECT p.hsn_code, p.gst_rate,
                           SUM(oi.taxable_value) - COALESCE(SUM(ri.refund_taxable_value), 0) as taxable,
                           (SUM(oi.cgst_amount) + SUM(oi.sgst_amount) + SUM(oi.igst_amount)) - COALESCE(SUM(ri.refund_cgst) + SUM(ri.refund_sgst) + SUM(ri.refund_igst), 0) as gst
                    FROM order_items oi
                    JOIN products p ON oi.product_id = p.id
                    JOIN orders o ON oi.order_id = o.id
                    LEFT JOIN (
                        SELECT ri.product_id, orr.order_id, 
                               SUM(ri.refund_taxable_value) as refund_taxable_value,
                               SUM(ri.refund_cgst) as refund_cgst, SUM(ri.refund_sgst) as refund_sgst, SUM(ri.refund_igst) as refund_igst
                        FROM return_items ri JOIN order_returns orr ON ri.return_id = orr.id
                        WHERE orr.status = 'completed' GROUP BY ri.product_id, orr.order_id
                    ) ri ON ri.product_id = oi.product_id AND ri.order_id = oi.order_id
                    WHERE o.status != 'cancelled' AND DATE(o.order_date) BETWEEN %s AND %s
                    GROUP BY p.hsn_code, p.gst_rate
                """, (start_date, end_date))
                for i, row in enumerate(cursor.fetchall(), 1):
                    writer.writerow([i, row['hsn_code'], f"{row['gst_rate']}%", f"{row['taxable']:.2f}", f"{row['gst']:.2f}"])

            elif report_id == 'profit_loss':
                writer.writerow(["Particulars", "Amount (Rs.)"])
                cursor.execute("SELECT SUM(total_amount) as revenue FROM orders WHERE DATE(order_date) BETWEEN %s AND %s AND status != 'cancelled'", (start_date, end_date))
                revenue = float(cursor.fetchone()['revenue'] or 0)
                cursor.execute("""
                    SELECT SUM(oi.quantity * COALESCE(p.material_cost,0)) as material_cost
                    FROM order_items oi JOIN products p ON oi.product_id = p.id
                    JOIN orders o ON oi.order_id = o.id
                    WHERE DATE(o.order_date) BETWEEN %s AND %s AND o.status != 'cancelled'
                """, (start_date, end_date))
                material = float(cursor.fetchone()['material_cost'] or 0)
                gross = revenue - material
                overhead = revenue * 0.08
                net = gross - overhead
                writer.writerow(["Total Revenue", f"{revenue:.2f}"])
                writer.writerow(["Less: Material Cost", f"({material:.2f})"])
                writer.writerow(["Gross Profit", f"{gross:.2f}"])
                writer.writerow(["Less: Operational Overhead (8%)", f"({overhead:.2f})"])
                writer.writerow(["Net Profit", f"{net:.2f}"])

            elif report_id == 'balance_sheet':
                writer.writerow(["Item", "Amount (Rs.)"])
                cursor.execute("SELECT COALESCE(SUM(stock_quantity * price),0) as inventory FROM products")
                inv = float(cursor.fetchone()['inventory'] or 0)
                cursor.execute("SELECT COALESCE(SUM(total_amount),0) as receivables FROM orders WHERE status NOT IN ('cancelled','delivered','refunded')")
                rec = float(cursor.fetchone()['receivables'] or 0)
                cursor.execute("SELECT COALESCE(SUM(total_amount),0) as cash FROM orders WHERE status IN ('delivered','completed')")
                cash = float(cursor.fetchone()['cash'] or 0)
                total_assets = inv + rec + cash
                writer.writerow(["ASSETS", ""])
                writer.writerow(["Inventory Value", f"{inv:.2f}"])
                writer.writerow(["Accounts Receivable", f"{rec:.2f}"])
                writer.writerow(["Cash & Bank", f"{cash:.2f}"])
                writer.writerow(["Total Assets", f"{total_assets:.2f}"])
                writer.writerow(["", ""])
                writer.writerow(["LIABILITIES & EQUITY", ""])
                writer.writerow(["GST Payable (estimated)", "0.00"])
                writer.writerow(["Owner's Equity", f"{total_assets:.2f}"])
                writer.writerow(["Total Liabilities & Equity", f"{total_assets:.2f}"])

            else:
                writer.writerow(["Report ID not supported for CSV", report_id])

        buffer = BytesIO()
        # Add UTF-8 BOM for Excel compatibility
        buffer.write(b'\xef\xbb\xbf')
        buffer.write(output.getvalue().encode('utf-8'))
        buffer.seek(0)
        return buffer, None

    except Exception as e:
        return None, str(e)